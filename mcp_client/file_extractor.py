"""
File Content Extractor
Extracts text content from various file types (PDF, DOCX, TXT).
"""

import logging
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)


def extract_text_from_file(content: bytes, filename: str) -> Optional[str]:
    """
    Extract text content from a file.
    
    Args:
        content: File content as bytes
        filename: Original filename (used to determine type)
    
    Returns:
        Extracted text or None if extraction fails
    """
    ext = Path(filename).suffix.lower()
    
    try:
        if ext == '.txt':
            return extract_from_txt(content)
        elif ext == '.pdf':
            return extract_from_pdf(content)
        elif ext in ['.docx', '.doc']:
            return extract_from_docx(content)
        elif ext in ['.xlsx', '.xls']:
            return extract_from_excel(content)
        elif ext in ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp']:
            return extract_from_image(content, filename)
        else:
            # Try as text
            try:
                return content.decode('utf-8')
            except:
                return f"[Unsupported file type: {ext}]"
    except Exception as e:
        logger.error(f"Extraction failed for {filename}: {e}")
        return f"[Error extracting from {filename}: {str(e)}]"


def extract_from_txt(content: bytes) -> str:
    """Extract text from TXT file."""
    try:
        return content.decode('utf-8')
    except UnicodeDecodeError:
        return content.decode('latin-1')


def extract_from_pdf(content: bytes) -> str:
    """Extract text from PDF file."""
    try:
        from pypdf import PdfReader
        import io
        
        reader = PdfReader(io.BytesIO(content))
        text_parts = []
        
        for page in reader.pages:
            text = page.extract_text()
            if text:
                text_parts.append(text)
        
        return "\n\n".join(text_parts)
    except ImportError:
        return "[PDF extraction requires pypdf package]"
    except Exception as e:
        return f"[PDF extraction error: {str(e)}]"


def extract_from_docx(content: bytes) -> str:
    """Extract text from DOCX file."""
    try:
        from docx import Document
        import io
        
        doc = Document(io.BytesIO(content))
        text_parts = []
        
        for para in doc.paragraphs:
            if para.text.strip():
                text_parts.append(para.text)
        
        return "\n\n".join(text_parts)
    except ImportError:
        return "[DOCX extraction requires python-docx package]"
    except Exception as e:
        return f"[DOCX extraction error: {str(e)}]"


def extract_from_excel(content: bytes) -> str:
    """Extract text from Excel file."""
    try:
        from openpyxl import load_workbook
        import io
        
        wb = load_workbook(io.BytesIO(content), read_only=True)
        text_parts = []
        
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            text_parts.append(f"[Sheet: {sheet}]")
            for row in ws.iter_rows(values_only=True):
                row_text = " | ".join(str(cell) for cell in row if cell is not None)
                if row_text.strip():
                    text_parts.append(row_text)
        
        return "\n".join(text_parts)
    except ImportError:
        return "[Excel extraction requires openpyxl package]"
    except Exception as e:
        return f"[Excel extraction error: {str(e)}]"


def extract_from_image(content: bytes, filename: str) -> str:
    """
    Extract information from image file.
    Uses Pillow for image metadata and basic description.
    """
    try:
        from PIL import Image
        import io
        
        img = Image.open(io.BytesIO(content))
        
        # Get image metadata
        width, height = img.size
        mode = img.mode
        format_type = img.format or "Unknown"
        
        # Build description
        description = f"[Image: {filename}]\n"
        description += f"Format: {format_type}\n"
        description += f"Size: {width}x{height} pixels\n"
        description += f"Mode: {mode}\n"
        
        # Check for EXIF data
        if hasattr(img, '_getexif') and img._getexif():
            exif = img._getexif()
            if exif:
                description += "Contains EXIF metadata\n"
        
        description += "\n[Image uploaded - visual content will be analyzed by the system]"
        
        return description
        
    except ImportError:
        return f"[Image {filename}: Pillow package required for image processing]"
    except Exception as e:
        return f"[Image {filename}: {str(e)}]"

